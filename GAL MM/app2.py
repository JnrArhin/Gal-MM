from flask import Flask, request, jsonify, render_template, redirect, url_for, session
import openpyxl
import os
import time
from flask_bcrypt import Bcrypt
from flask_session import Session

app = Flask(__name__)
app.secret_key = 'your_secret_key'  # Change this to a random secret key
app.config['SESSION_TYPE'] = 'filesystem'
Session(app)
bcrypt = Bcrypt(app)

# In-memory user storage (for demonstration purposes)
users = {}

# Path to your Excel workbook
EXCEL_FILE_PATH = os.path.expanduser("~/Desktop/mining_data.xlsx")

# Function to save data to the specified sheet
def save_to_excel(sheet_name, data):
    while True:
        try:
            if os.path.exists(EXCEL_FILE_PATH):
                workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
            else:
                workbook = openpyxl.Workbook()

            if sheet_name not in workbook.sheetnames:
                sheet = workbook.create_sheet(sheet_name)
                headers = list(data.keys())
                sheet.append(headers)
            else:
                sheet = workbook[sheet_name]

            sheet.append(list(data.values()))
            workbook.save(EXCEL_FILE_PATH)
            print(f"Data saved to {sheet_name}: {data}")
            break
        except PermissionError:
            print("Workbook is open. Waiting to try again...")
            time.sleep(5)
        except Exception as e:
            print(f"Error saving to Excel: {e}")
            break

@app.route('/')
def index():
    return render_template('index3.html')

@app.route('/register', methods=['POST'])
def register():
    username = request.form['username']
    password = request.form['password']
    hashed_password = bcrypt.generate_password_hash(password).decode('utf-8')
    users[username] = hashed_password  # Store the user
    return redirect(url_for('login'))  # Redirect to login after registration

@app.route('/login', methods=['POST'])
def login():
    username = request.form['username']
    password = request.form['password']
    if username in users and bcrypt.check_password_hash(users[username], password):
        session['username'] = username  # Store username in session
        return render_template('index3.html', active_tab='dashboard')  # Render index3.html with dashboard active
    return "Invalid credentials", 401  # Return error for invalid credentials

@app.route('/logout')
def logout():
    session.pop('username', None)  # Remove username from session
    return redirect(url_for('index'))  # Redirect to home page

@app.route('/submit_mining_material_data', methods=['POST'])
def submit_mining_material_data():
    mining_date = request.form['mining-date']
    mining_shift = request.form['mining-shift']
    ore_truck_count = request.form['ore-truck-count']
    waste_truck_count = request.form['waste-truck-count']

    data = {
        'Date': mining_date,
        'Shift': mining_shift,
        'Total Ore Truck Count': ore_truck_count,
        'Total Waste Truck Count': waste_truck_count
    }

    try:
        save_to_excel('Mining Materials', data)
        return jsonify({"message": "Mining materials submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_equipment_stats_data', methods=['POST'])
def submit_equipment_stats_data():
    equipment_id = request.form['equipment-id']
    start_hour_meter = request.form['start-hour-meter']
    end_hour_meter = request.form['end-hour-meter']

    data = {
        'Equipment ID': equipment_id,
        'Start Hour Meter': start_hour_meter,
        'End Hour Meter': end_hour_meter
    }

    try:
        save_to_excel('Equipment Statistics', data)
        return jsonify({"message": "Equipment statistics submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_exploration_geology_metrics', methods=['POST'])
def submit_exploration_geology_metrics():
    field_selection = request.form['field-selection']
    supervisor = request.form['supervisor']

    data = {
        'Field Selection': field_selection,
        'Supervisor': supervisor
    }

    # Add specific data based on the selected field
    if field_selection == 'mapping':
        data.update({
            'Map Date': request.form['map-date'],
            'Map Shift': request.form['map-shift'],
            'Map Scale': request.form['map-scale'],
            'Traverse Distance': request.form['traverse-distance'],
            'Area Covered': request.form['area-covered'],
        })
    elif field_selection == 'geophysical':
        data.update({
            'Survey Date': request.form['survey-date'],
            'Survey Shift': request.form['survey-shift'],
            'Survey Method': request.form['survey-method'],
            'Survey Area': request.form['survey-area'],
            'Equipment Used': request.form['equipment-used'],
            'Anomalies Detected': request.form['anomalies-detected'],
            'Interpretation': request.form['interpretation'],
        })
    elif field_selection == 'geochemical':
        data.update({
            'Geochem Date': request.form['geochem-date'],
            'Geochem Shift': request.form['geochem-shift'],
            'Sample ID Range': request.form['sample-id-range'],
            'Number of Samples': request.form['num-samples'],
            'Sampling Geologist': request.form['sampling-geologist'],
            'Sampling Method': request.form['sampling-method'],
            'Laboratory': request.form['laboratory'],
        })
    elif field_selection == 'trenching':
        data.update({
            'Trench Date': request.form['trench-date'],
            'Trench Shift': request.form['trench-shift'],
            'Trench ID': request.form['trench-id'],
            'Trench Length': request.form['trench-length'],
            'Trench Width': request.form['trench-width'],
            'Sample Results': request.form['sample-results'],
            'Geological Description': request.form['geological-description'],
            'Channel Sampling': request.form['channel-sampling'],
        })
    elif field_selection == 'drilling':
        data.update({
            'Drill Date': request.form['drill-date'],
            'Drill Shift': request.form['drill-shift'],
            'Drill Hole ID': request.form['drill-hole-id'],
            'Drill Depth': request.form['drill-depth'],
            'Core Recovery': request.form['core-recovery'],
            'Gold Assay': request.form['gold-assay'],
            'Lithology': request.form['lithology'],
            'Structural Analysis': request.form['structural-analysis'],
            'Downhole Logs': request.form['downhole-logs'],
            'Sampling Technique': request.form['sampling-technique'],
        })

    try:
        save_to_excel('Exploration Geology Metrics', data)
        return jsonify({"message": "Exploration geology metrics submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# Hazard Reports Section
@app.route('/submit_hazard_report_data', methods=['POST'])
def submit_hazard_report_data():
    data = request.form.to_dict()
    try:
        save_to_excel('Hazard Reports', data)
        return jsonify({"message": "Hazard report data saved successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# Maintenance Data Section
@app.route('/submit_maintenance_data', methods=['POST'])
def submit_maintenance_data():
    data = request.form.to_dict()
    try:
        save_to_excel('Maintenance Data', data)
        return jsonify({"message": "Maintenance data saved successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_crushing_data', methods=['POST'])
def submit_crushing_data():
    data = request.form.to_dict()
    try:
        save_to_excel('Crushing Data', data)
        return jsonify({"message": "Crushing data submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_solution_management_data', methods=['POST'])
def submit_solution_management_data():
    data = request.form.to_dict()
    try:
        save_to_excel('Solution Management Data', data)
        return jsonify({"message": "Solution management data submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_geophysics_data', methods=['POST'])
def submit_geophysics_data():
    data = request.form.to_dict()
    try:
        save_to_excel('Geophysical Data', data)
        return jsonify({"message": "Geophysical data submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_geochemical_data', methods=['POST'])
def submit_geochemical_data():
    data = request.form.to_dict()
    try:
        save_to_excel('Geochemical Data', data)
        return jsonify({"message": "Geochemical data submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_trenching_data', methods=['POST'])
def submit_trenching_data():
    data = request.form.to_dict()
    try:
        save_to_excel('Trenching Data', data)
        return jsonify({"message": "Trenching data submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_drilling_data', methods=['POST'])
def submit_drilling_data():
    data = request.form.to_dict()
    try:
        save_to_excel('Drilling Data', data)
        return jsonify({"message": "Drilling data submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_incident_report_data', methods=['POST'])
def submit_incident_report_data():
    data = request.form.to_dict()
    try:
        save_to_excel('Incident Reports', data)
        return jsonify({"message": "Incident report submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_monitoring_data', methods=['POST'])
def submit_monitoring_data():
    data = request.form.to_dict()
    try:
        save_to_excel('Air and Noise Monitoring', data)
        return jsonify({"message": "Monitoring data submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_water_sample_data', methods=['POST'])
def submit_water_sample_data():
    data = request.form.to_dict()
    try:
        save_to_excel('Water Samples', data)
        return jsonify({"message": "Water sample data submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_inspection_data', methods=['POST'])
def submit_inspection_data():
    data = request.form.to_dict()
    try:
        save_to_excel('Site Inspections', data)
        return jsonify({"message": "Inspection data submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_complaint_data', methods=['POST'])
def submit_complaint_data():
    data = request.form.to_dict()
    try:
        save_to_excel('Complaints', data)
        return jsonify({"message": "Complaint submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_requests_data', methods=['POST'])
def submit_requests_data():
    data = request.form.to_dict()
    try:
        save_to_excel('Requests', data)
        return jsonify({"message": "Request submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_stakeholder_engagement_data', methods=['POST'])
def submit_stakeholder_engagement_data():
    data = request.form.to_dict()
    try:
        save_to_excel('Stakeholder Engagement', data)
        return jsonify({"message": "Stakeholder engagement data submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_assessment_data', methods=['POST'])
def submit_assessment_data():
    data = request.form.to_dict()
    try:
        save_to_excel('Land and Crop Assessments', data)
        return jsonify({"message": "Land and crop assessment data submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', debug=True)  # Allow access from other computers