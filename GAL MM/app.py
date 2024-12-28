from flask import Flask, request, jsonify, render_template, redirect, url_for, session
from flask_sqlalchemy import SQLAlchemy
from flask_session import Session
import os
import openpyxl
import time

app = Flask(__name__)
app.secret_key = 'your_secret_key'  # Change this to a random secret key
app.config['SESSION_TYPE'] = 'filesystem'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///users.db'  # SQLite database
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Initialize the database after the app is created
db = SQLAlchemy(app)
Session(app)

# Path to your Excel workbook
EXCEL_FILE_PATH = os.path.expanduser("~/Desktop/mining_data.xlsx")

# User model for authentication
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password = db.Column(db.String(120), nullable=False)

# Function to save data to the specified sheet
def save_to_excel(sheet_name, data):
    while True:
        try:
            if os.path.exists(EXCEL_FILE_PATH):
                workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
            else:
                workbook = openpyxl.Workbook()

            if sheet_name not in workbook.sheetnames:
                sheet = workbook.create_sheet(sheet_name)
                headers = list(data.keys())
                sheet.append(headers)
            else:
                sheet = workbook[sheet_name]

            sheet.append(list(data.values()))
            workbook.save(EXCEL_FILE_PATH)
            print(f"Data saved to {sheet_name}: {data}")
            break
        except PermissionError:
            print("Workbook is open. Waiting to try again...")
            time.sleep(5)
        except Exception as e:
            print(f"Error saving to Excel: {e}")
            break

# Function to check for duplicates
def check_for_duplicates(sheet_name, date, shift, equipment_id=None):
    if os.path.exists(EXCEL_FILE_PATH):
        workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header
                if row[0] == date and row[1] == shift:  # Check only date and shift
                    return True
    return False

@app.route('/')
def index():
    return redirect(url_for('login'))  # Redirect to the login page

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        pass
        username = request.form['username']
        password = request.form['password']
        user = User.query.filter_by(username=username, password=password).first()
        if user:
            session['user_id'] = user.id
            print(f"User {username} logged in successfully.")
            return redirect(url_for('dashboard'))  # Redirect to dashboard after login
        else:
            return "Invalid credentials", 401
    return render_template('login.html')  # Render the login page

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        # Check if the username already exists
        existing_user = User.query.filter_by(username=username).first()
        if existing_user:
            return "Username already exists. Please choose a different one.", 400  # Bad request

        new_user = User(username=username, password=password)
        db.session.add(new_user)
        db.session.commit()
        return redirect(url_for('login'))  # Redirect to login after registration
    return render_template('register.html')  # Render the registration page

@app.route('/dashboard')
def dashboard():
    if 'user_id' not in session:
        print("User not authenticated, redirecting to login.")
        return redirect(url_for('login'))  # Redirect to login if not authenticated
    print("Rendering dashboard.")
    return render_template('index3.html')  # Render the dashboard page

@app.route('/submit_mining_material_data', methods=['POST'])
def submit_mining_material_data():
    mining_date = request.form['mining-date']
    mining_shift = request.form['mining-shift']
    user_action = request.form.get('user_action', 'none')  # "continue", "overwrite", or "cancel"

    # Check for duplicates
    is_duplicate = check_for_duplicates('Mining Materials', mining_date, mining_shift)
    if is_duplicate and user_action == 'none':
        return jsonify({
            "error": "Duplicate entry found.",
            "options": ["continue", "overwrite", "cancel"]
        }), 409  # Conflict status code

    data = {
        'Date': mining_date,
        'Shift': mining_shift,
        'Total Ore Truck Count': request.form ['ore-truck-count'],
        'Total Waste Truck Count': request.form ['waste-truck-count']
    }

    try:
        save_to_excel('Mining Materials', data)
        return jsonify({"message": "Mining materials submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_equipment_stats_data', methods=['POST'])
def submit_equipment_stats_data():
    mining_date = request.form['mining-date']
    mining_shift = request.form['mining-shift']
    equipment_id = request.form['equipment-id']
    user_action = request.form.get('user_action', 'none')  # "continue", "overwrite", or "cancel"

    # Check for duplicates
    is_duplicate = check_for_duplicates('Equipment Statistics', mining_date, mining_shift, equipment_id)
    if is_duplicate and user_action == 'none':
        return jsonify({
            "error": "Duplicate entry found.",
            "options": ["continue", "overwrite", "cancel"]
        }), 409  # Conflict status code

    if is_duplicate and user_action == 'cancel':
        return jsonify({"message": "Action cancelled by the user."}), 200

    if is_duplicate and user_action == 'continue':
        # Proceed without overwriting
        return jsonify({"message": "Data submission continued without changes."}), 200

    # If no duplicate or user chooses to overwrite
    data = {
        'Mining Date': mining_date,
        'Mining Shift': mining_shift,
        'Equipment ID': equipment_id,
        'Start Hour Meter': request.form['start-hour-meter'],
        'End Hour Meter': request.form['end-hour-meter'],
        'Equipment Run Hours': request.form['equipment-run-hours'],
        'Fuel Recieved': request.form['fuel-recieved'],
        'Mining Operation Start Time': request.form['mining-operations-start-time'],
        'Mining Operation End Time': request.form['mining-operations-end-time'],
        'Total Production Hours': request.form['total-production-hours'],
        'Total Downtime Hours': request.form['total-downtime-hours'],
        'Total Tramming Hours': request.form['total-tramming-hours'],
        'Operational GSA Hours': request.form['operational-gsa-hours'],
        'Other GSA Hours': request.form['other-gsa-hours'],
        'Equipment Standby': request.form['equipment-standby'],
        'Comments': request.form['comments'],
        'Operator Name': request.form['operator-name'],
        'Supervisor Name': request.form['Supervisor Name'],
    }

    try:
        save_to_excel('Equipment Statistics', data)
        return jsonify({"message": "Equipment stats submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_exploration_geology_metrics', methods=['POST'])
def submit_exploration_geology_metrics():
    field_selection = request.form['field-selection']
    supervisor = request.form['supervisor']

    data = {
        'Field Selection': field_selection,
        'Supervisor': supervisor
    }

    # Add specific data based on the selected field
    if field_selection == 'mapping':
        data.update({
            'Map Date': request.form['map-date'],
            'Map Shift': request.form['map-shift'],
            'Map Scale': request.form['map-scale'],
            'Traverse Distance': request.form['traverse-distance'],
            'Area Covered': request.form['area-covered'],
        })
    elif field_selection == 'geophysical':
        data.update({
            'Survey Date': request.form['survey-date'],
            'Survey Shift': request.form['survey-shift'],
            'Survey Method': request.form['survey-method'],
            'Survey Area': request.form['survey-area'],
            'Equipment Used': request.form['equipment-used'],
            'Anomalies Detected': request.form['anomalies-detected'],
            'Interpretation': request.form['interpretation'],
        })
    elif field_selection == 'geochemical':
        data.update({
            'Geochem Date': request.form['geochem-date'],
            'Geochem Shift': request.form['geochem-shift'],
            'Sample ID Range': request.form['sample-id-range'],
            'Number of Samples': request.form['num-samples'],
            'Sampling Geologist': request.form['sampling-geologist'],
            'Sampling Method': request.form['sampling-method'],
            'Laboratory': request.form['laboratory'],
        })
    elif field_selection == 'trenching':
        data.update({
            'Trench Date': request.form['trench-date'],
            'Trench Shift': request.form['trench-shift'],
            'Trench ID': request.form['trench-id'],
            'Trench Length': request.form['trench-length'],
            'Trench Width': request.form['trench-width'],
            'Sample Results': request.form['sample-results'],
            'Geological Description': request.form['geological-description'],
            'Channel Sampling': request.form['channel-sampling'],
        })
    elif field_selection == 'drilling':
        data.update({
            'Drilling Date': request.form['drilling-date'],
            'Drilling Shift': request.form['drilling-shift'],
            'Drill Hole ID': request.form['drill-hole-id'],
            'Depth': request.form['depth'],
            'Core Recovery': request.form['core-recovery'],
            'Lithology': request.form['lithology'],
            'Structural Analysis': request.form['structural-analysis'],
            'Downhole Logs': request.form['downhole-logs'],
            'Sampling Technique': request.form['sampling-technique'],
        })

    try:
        save_to_excel('Exploration Geology Metrics', data)
        return jsonify({"message": "Exploration geology metrics submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# Hazard Reports Section
@app.route('/submit_hazard_report_data', methods=['POST'])
def submit_hazard_report_data():
    data = request.form.to_dict()
    try:
        save_to_excel('Hazard Reports', data)
        return jsonify({"message": "Hazard report data saved successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# Maintenance Data Section
@app.route('/submit_maintenance_data', methods=['POST'])
def submit_maintenance_data():
    data = request.form.to_dict()
    try:
        save_to_excel('Maintenance Data', data)
        return jsonify({"message": "Maintenance data saved successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_crushing_data', methods=['POST'])
def submit_crushing_data():
    data = request.form.to_dict()
    try:
        save_to_excel('Crushing Data', data)
        return jsonify({"message": "Crushing data submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_solution_management_data', methods=['POST'])
def submit_solution_management_data():
    data = request.form.to_dict()
    try:
        save_to_excel('Solution Management Data', data)
        return jsonify({"message": "Solution management data submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_geophysics_data', methods=['POST'])
def submit_geophysics_data():
    data = request.form.to_dict()
    try:
        save_to_excel('Geophysical Data', data)
        return jsonify({"message": "Geophysical data submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_geochemical_data', methods=['POST'])
def submit_geochemical_data():
    data = request.form.to_dict()
    try:
        save_to_excel('Geochemical Data', data)
        return jsonify({"message": "Geochemical data submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_trenching_data', methods=['POST'])
def submit_trenching_data():
    data = request.form.to_dict()
    try:
        save_to_excel('Trenching Data', data)
        return jsonify({"message": "Trenching data submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_drilling_data', methods=['POST'])
def submit_drilling_data():
    data = request.form.to_dict()
    try:
        save_to_excel('Drilling Data', data)
        return jsonify({"message": "Drilling data submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_incident_report_data', methods=['POST'])
def submit_incident_report_data():
    data = request.form.to_dict()
    try:
        save_to_excel('Incident Reports', data)
        return jsonify({"message": "Incident report submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_monitoring_data', methods=['POST'])
def submit_monitoring_data():
    data = request.form.to_dict()
    try:
        save_to_excel('Air and Noise Monitoring', data)
        return jsonify({"message": "Monitoring data submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_water_sample_data', methods=['POST'])
def submit_water_sample_data():
    data = request.form.to_dict()
    try:
        save_to_excel('Water Samples', data)
        return jsonify({"message": "Water sample data submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_inspection_data', methods=['POST'])
def submit_inspection_data():
    data = request.form.to_dict()
    try:
        save_to_excel('Site Inspections', data)
        return jsonify({"message": "Inspection data submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_complaint_data', methods=['POST'])
def submit_complaint_data():
    data = request.form.to_dict()
    try:
        save_to_excel('Complaints', data)
        return jsonify({"message": "Complaint submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_requests_data', methods=['POST'])
def submit_requests_data():
    data = request.form.to_dict()
    try:
        save_to_excel('Requests', data)
        return jsonify({"message": "Request submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_stakeholder_engagement_data', methods=['POST'])
def submit_stakeholder_engagement_data():
    data = request.form.to_dict()
    try:
        save_to_excel('Stakeholder Engagement', data)
        return jsonify({"message": "Stakeholder engagement data submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_assessment_data', methods=['POST'])
def submit_assessment_data():
    data = request.form.to_dict()
    try:
        save_to_excel('Land and Crop Assessments', data)
        return jsonify({"message": "Land and crop assessment data submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    with app.app_context():  # Ensure the application context is set
        db.create_all()  # Create tables if they don't exist
    app.run(host='0.0.0.0', debug=True)  # Allow access from other computers